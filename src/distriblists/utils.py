import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment, Border, Side

from django.utils.translation import ugettext_lazy as _

from accounts.models import User
from distriblists.models import DistributionList
from distriblists.forms import DistributionListForm


header_alignment = Alignment(
    horizontal='center',
    textRotation=45,
)

role_alignment = Alignment(
    horizontal='center'
)
cell_border = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin'),
    vertical=Side(style='thin')
)


def import_review_members(filepath, category):
    """Import review members from an excel file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Extracts the user list from the header row
    emails, user_ids = _extract_users(ws)

    max_col = len(user_ids) + 1  # Don't use ws.max_column, it's not reliable
    rows = ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col)
    results = []
    for idx, row in enumerate(rows):
        result = _import_review_members(row, emails, user_ids, category)
        result['line'] = idx + 2
        results.append(result)

    return results


def _import_review_members(row, emails, user_ids, category):
    """Saves a review member list for a single row."""
    errors = []

    key = row[0].value
    try:
        instance = category.document_class().objects \
            .filter(document__category=category) \
            .select_related('latest_revision') \
            .get(document__document_key=key)
    except:  # noqa
        instance = None
        errors.append(_('Document {} does not exist').format(key))

    if instance.latest_revision.is_under_review():
        errors.append(_('This document is under review'))

    # Extract user roles from xls
    reviewers = []
    leader = None
    approver = None
    for idx, cell in enumerate(row[1:]):
        role = cell.value
        if role:
            user_id = user_ids[idx]
            if user_id is None:
                errors.append('Unknown user "{}"'.format(emails[idx]))

            if role == 'R':
                reviewers.append(user_id)
            elif role == 'L':
                leader = user_id
            elif role == 'A':
                approver = user_id
            else:
                errors.append('Unknown role "{}"'.format(role))

    if instance and not errors:
        rev = instance.latest_revision
        if leader:
            rev.leader_id = leader

        if approver:
            rev.approver_id = approver
        rev.reviewers.clear()
        rev.reviewers.add(*reviewers)
        rev.save()

    return {
        'document_key': key,
        'success': not errors,
        'errors': errors,
    }


def export_review_members(category):
    """Export members of the review for all documents in the category."""
    documents = category.document_class().objects \
        .filter(document__category=category) \
        .select_related(
            'document',
            'latest_revision',
            'latest_revision__leader',
            'latest_revision__approver') \
        .prefetch_related('latest_revision__reviewers') \
        .order_by('document__document_number')
    users_qs = User.objects.filter(categories=category).order_by('email')
    all_users = list(users_qs)

    wb = openpyxl.Workbook()
    ws = wb.active

    _export_header(ws, all_users)
    for idx, doc in enumerate(documents):
        _export_members(ws, idx, doc, all_users)

    return save_virtual_workbook(wb)


def _export_members(ws, idx, doc, all_users):
    """Add a single line to the exported file."""
    line = idx + 2
    ws.cell(row=line, column=1).value = doc.document.document_number

    rev = doc.latest_revision

    if rev.leader:
        _export_role(ws, line, all_users, rev.leader, 'L')

    if rev.approver:
        _export_role(ws, line, all_users, rev.approver, 'A')

    for reviewer in rev.reviewers.all():
        _export_role(ws, line, all_users, reviewer, 'R')

    # Set borders for all cells
    for column in range(1, len(all_users) + 2):
        ws.cell(row=line, column=column).border = cell_border


def export_lists(category):
    """Export distribution lists in a single category as xlsx file."""
    lists = DistributionList.objects \
        .filter(categories=category) \
        .select_related('leader', 'approver') \
        .prefetch_related('reviewers')
    users_qs = User.objects.filter(categories=category).order_by('email')
    all_users = list(users_qs)

    wb = openpyxl.Workbook()
    ws = wb.active

    _export_header(ws, all_users)

    for idx, dlist in enumerate(lists):
        _export_list(ws, idx, dlist, all_users)

    return save_virtual_workbook(wb)


def _export_header(ws, all_users):
    """Add header row to exported file."""
    header_row = next(ws.get_squared_range(
        min_col=2,
        min_row=1,
        max_col=len(all_users) + 1,
        max_row=1))
    for idx, cell in enumerate(header_row):
        cell.value = all_users[idx].email
        cell.alignment = header_alignment
        cell.border = cell_border


def _export_list(ws, idx, dlist, all_users):
    """Add a single line to the exported file."""
    line = idx + 2
    ws.cell(row=line, column=1).value = dlist.name

    _export_role(ws, line, all_users, dlist.leader, 'L')

    if dlist.approver:
        _export_role(ws, line, all_users, dlist.approver, 'A')

    for reviewer in dlist.reviewers.all():
        _export_role(ws, line, all_users, reviewer, 'R')

    # Set borders for all cells
    for column in range(1, len(all_users) + 2):
        ws.cell(row=line, column=column).border = cell_border


def _export_role(ws, line, all_users, user, role):
    """Set a single cell in exported file."""
    user_index = all_users.index(user)
    cell = ws.cell(row=line, column=user_index + 2)
    cell.value = role
    cell.alignment = role_alignment


def import_lists(filepath, category):
    """Import distribution lists from an excel file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Extracts the user list from the header row
    emails, user_ids = _extract_users(ws)

    max_col = len(user_ids) + 1  # Don't use ws.max_column, it's not reliable
    rows = ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col)
    results = []
    for idx, row in enumerate(rows):
        result = _import_list(row, emails, user_ids, category)
        result['line'] = idx + 2
        results.append(result)

    return results


def _extract_users(ws):
    """Extract the xls header, i.e the list of email users.

    We cannot rely on the worksheet's dimensions and, say, extract the first
    row until the latest column because this value is sometimes off, especially
    with documents created with LibreOffice.

    Return the list of user ids.

    """
    emails = []
    column_index = 2
    user_cell = ws.cell(row=1, column=column_index)
    while user_cell.value:
        emails.append(user_cell.value)
        column_index += 1
        user_cell = ws.cell(row=1, column=column_index)

    qs = User.objects.filter(email__in=emails) \
        .values_list('email', 'id')
    user_dict = dict(qs)

    user_ids = [user_dict.get(email, None) for email in emails]
    return emails, user_ids


def _import_list(row, emails, user_ids, category):
    """Saves a distribution list for a single row."""
    errors = []

    # Fetch existing list if it exists
    list_name = row[0].value
    try:
        instance = DistributionList.objects.get(name=list_name)
        categories = list(instance.categories.all())
        categories.append(category)
        action = _('Update')
    except DistributionList.DoesNotExist:
        instance = None
        categories = [category]
        action = _('Create')

    # Extract user roles from xls
    reviewers = []
    leader = None
    approver = None
    for idx, cell in enumerate(row[1:]):
        role = cell.value
        if role:
            user_id = user_ids[idx]
            if user_id is None:
                errors.append('Unknown user "{}"'.format(emails[idx]))

            if role == 'R':
                reviewers.append(user_id)
            elif role == 'L':
                leader = user_id
            elif role == 'A':
                approver = user_id
            else:
                errors.append('Unknown role "{}"'.format(role))

    # Use the model form to validate and save the data
    data = {
        'name': list_name,
        'categories': [c.id for c in categories],
        'reviewers': reviewers,
        'leader': leader if leader else None,
        'approver': approver if approver else None,
    }
    form = DistributionListForm(data, instance=instance)
    if form.is_valid() and not errors:
        form.save()

    if not form.is_valid():
        form_errors = [field_errors[0] for field_errors in list(form.errors.values())]
        errors.append(*form_errors)

    return {
        'list_name': list_name,
        'action': action,
        'success': form.is_valid() and not errors,
        'errors': errors,
    }
